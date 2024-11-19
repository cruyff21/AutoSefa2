from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep

# Função para ler os dados do Excel
def ler_dados_do_excel(nome_arquivo):
    workbook = load_workbook(nome_arquivo)
    sheet = workbook.active
    dados = []

    # Itera pelas linhas a partir da segunda (pula o cabeçalho)
    for row in sheet.iter_rows(min_row=63, max_row=63):
        cpf = row[1].value  # CPF na coluna B
        senha = row[2].value  # Senha na coluna C
        inscricao = row[3].value  # Inscrição estadual na coluna D
        data_inicial = row[4].value  # Data inicial na coluna E
        data_final = row[5].value  # Data final na coluna F
        dados.append((row[0].row, cpf, senha, inscricao, data_inicial, data_final))  # Adiciona o número da linha
    return workbook, sheet, dados

# Função para atualizar o status na planilha
def atualizar_status(sheet, linha, status):
    try:
        coluna_status = 7  # Coluna G para o status (mude se necessário)
        sheet.cell(row=linha, column=coluna_status, value=status)
    except Exception as e:
        print(f"Erro ao atualizar o status na linha {linha}: {e}")

# Configurar o WebDriver
def configurar_driver():
    options = Options()
    #options = webdriver.ChromeOptions()
    #options.add_argument("--headless=new")
    # options.add_experimental_option("detach", True)  # Mantém o navegador aberto
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.maximize_window()
    return driver

# Realizar login no site
def realizar_login(driver, cpf, senha):
    try:
        driver.get("https://app.sefa.pa.gov.br/pservicos/autenticacao?servico=aHR0cHM6Ly9hcHAuc2VmYS5wYS5nb3YuYnIvYXJxdWl2b3MtZG93bmxvYWQvIy8=")
        driver.find_element(By.XPATH, '//*[@id="cpf"]').send_keys(cpf)
        driver.find_element(By.XPATH, '//*[@id="senha"]').send_keys(senha)
        login_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="btn-entrar"]'))
        )
        login_button.click()
        print("Login realizado")
    except Exception as e:
        print(f"Erro durante o login para CPF {cpf}: {e}")
        raise e

# Encontrar e clicar na inscrição correspondente
def clicar_inscricao(driver, inscricao_excel):
    try:
        WebDriverWait(driver, 100).until(
            EC.presence_of_all_elements_located((By.XPATH, '//*[@class="primary-text ng-binding"]'))
        )
        inscricoes = driver.find_elements(By.XPATH, '//*[@class="primary-text ng-binding"]')
        for inscricao in inscricoes:
            if inscricao.text == str(inscricao_excel):
                print(f"Encontrado: {inscricao.text}")
                inscricao.click()
                return True
        print("Inscrição estadual não encontrada!")
        return False
    except Exception as e:
        print(f"Erro ao localizar inscrição estadual: {e}")
        return False

# Realizar as interações adicionais no site
def interagir_com_site(driver, data_inicial, data_final):
    try:
        data_inicial_formatada = data_inicial.strftime('%d/%m/%Y')
        data_final_formatada = data_final.strftime('%d/%m/%Y')

        # Aguardar e clicar na nova solicitação
        WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="content"]/div/div/div/div/async-loader/div/md-card/div[1]/div[1]/button[1]'))
        ).click()


        # Espera até que o campo (dropdown) esteja visível e clica nele
        
        tipo_de_arquivo = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="select_17"]'))
        )
        tipo_de_arquivo.click()  # Clica para abrir a lista de opções


        # Espera até que a opção desejada esteja visível e clica nela
        ct_e = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="select_option_23"]'))
        )
        ct_e.click()  # Clica na opção desejada

         # Preencher datas e realizar outras ações conforme necessário
        WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="input_35"]'))
        ).send_keys(data_inicial_formatada)

        WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="input_37"]'))
        ).send_keys(data_final_formatada)
        

        entrada = driver.find_element(By.XPATH, '//*[@id="tab-content-32"]/div/md-content/div/md-input-container[3]/div[2]/div[1]/div/md-checkbox')
        driver.execute_script("arguments[0].scrollIntoView();", entrada)
        # Espera e clica no checkbox
        entrada.click()
        
        adicionar = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="content"]/div/div/div/form/async-loader/div/md-card/div/button[1]'))
        )
        adicionar.click()
        print(f"CT-e adicionado entre {data_inicial_formatada} e {data_final_formatada}")
        

        # Espera até que o campo (dropdown) esteja visível e clica nele
        tipo_de_arquivo = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="select_17"]'))
        )
        tipo_de_arquivo.click()  # Clica para abrir a lista de opções

        # Espera até que a opção desejada esteja visível e clica nela
        nf_e_saida = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="select_option_29"]'))
        )
        nf_e_saida.click()  # Clica na opção desejada

        # Preencher datas e realizar outras ações conforme necessário
        WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="input_35"]'))
        ).send_keys(data_inicial_formatada)

        WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="input_37"]'))
        ).send_keys(data_final_formatada)

        saida = driver.find_element(By.XPATH, '//*[@id="tab-content-32"]/div/md-content/div/md-input-container[3]/div[2]/div[2]/div/md-checkbox')
        driver.execute_script("arguments[0].scrollIntoView();", saida)
        # Espera e clica no checkbox
        saida.click()

        # Espera até que o elemento esteja visível e interativo
        adicionar = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="content"]/div/div/div/form/async-loader/div/md-card/div/button[1]'))
        )
        adicionar.click()
        print(f"NF-e de saida adicionado entre {data_inicial_formatada} e {data_final_formatada}")

        # Espera até que o campo (dropdown) esteja visível e clica nele
        tipo_de_arquivo = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="select_17"]'))
        )
        tipo_de_arquivo.click()  # Clica para abrir a lista de opções

        # Espera até que a opção desejada esteja visível e clica nela
        nfc_e = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="select_option_30"]'))
        )
        nfc_e.click()  # Clica na opção desejada

        # Preencher datas e realizar outras ações conforme necessário
        WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="input_35"]'))
        ).send_keys(data_inicial_formatada)

        WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="input_37"]'))
        ).send_keys(data_final_formatada)

        # Espera até que o elemento esteja visível e interativo
        adicionar = WebDriverWait(driver, 1000).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="content"]/div/div/div/form/async-loader/div/md-card/div/button[1]'))
        )
        adicionar.click()
        print(f"NFC-e adicionado entre {data_inicial_formatada} e {data_final_formatada}")

        # Aguarda até que o elemento esteja presente no DOM e seja visível
        # Aguarda o checkbox estar presente no DOM
        checkbox_container = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="content"]/div/div/div/form/async-loader/div/async-loader/div/md-card/md-table-container/table/thead[1]/tr/th[1]/md-checkbox/div'))
        )

        # Certifica-se de que o elemento está visível e rola até ele
        driver.execute_script("arguments[0].scrollIntoView();", checkbox_container)

        # Clica no checkbox
        try:
            checkbox_container.click()
            print("Todos os arquivos selecionados")
        except Exception as e:
            print(f"Erro ao clicar no checkbox: {e}")


         # Aguarde até que o toast desapareça
        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element((By.CLASS_NAME, "md-toast-text"))
        )

        # Agora tente clicar no botão
        botao_enviar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@ng-click="ctrl.enviar()"]'))
        )
        try:
            botao_enviar.click()
            print("Arquivos enviados com sucesso!")
        except Exception as e:
            print(f"Erro ao clicar em enviar {e}")
        sleep(4)

        # Aqui vai o seu código para interagir com o site
        print("Interações realizadas com sucesso.")
    except Exception as e:
        print(f"Erro durante as interações: {e}")
        

# Integração geral
def main():
    arquivo_excel = "Empresas_sefa.xlsx"
    workbook, sheet, dados = ler_dados_do_excel(arquivo_excel)

    for linha, cpf, senha, inscricao, data_inicial, data_final in dados:
        driver = configurar_driver()
        try:
            realizar_login(driver, cpf, senha)
            encontrado = clicar_inscricao(driver, inscricao)
            if encontrado:
                interagir_com_site(driver, data_inicial, data_final)
                atualizar_status(sheet, linha, "CONCLUÍDO")
            else:
                atualizar_status(sheet, linha, "Inscrição não encontrada")
        except Exception as e:
            atualizar_status(sheet, linha, f"Erro: {e}")
        
            

    # Salvar as alterações no Excel
    workbook.save(arquivo_excel)
    print("Processo concluído e planilha atualizada.")

if __name__ == "__main__":
    main()
